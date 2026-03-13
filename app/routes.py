import io
import os
import threading
import time
from datetime import date
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, g, current_app
from .models import (
    get_all_clients, get_client, add_client, delete_client,
    get_clients_by_first_name,
    get_vl1, save_vl1, get_vl2, save_vl2, get_vl3, save_vl3,
    get_all_for_export, SW_QUESTIONS, calc_sw_scores,
    get_dashboard_data,
)
from . import get_coach_settings, save_coach_settings, GEMEENTEN
from .updates import UpdateError, get_update_status, start_self_update
from .version import APP_VERSION

bp = Blueprint('main', __name__)


@bp.before_request
def load_coach():
    """Make coach settings available in g for every request."""
    g.coach = get_coach_settings()


@bp.context_processor
def inject_coach():
    return {'coach': g.coach, 'app_version': APP_VERSION}


def _schedule_process_exit(delay_seconds=1.0):
    def _exit_later():
        time.sleep(delay_seconds)
        os._exit(0)

    threading.Thread(target=_exit_later, daemon=True).start()


# ── Coach setup / wijzigen ────────────────────────────────────────────────────

@bp.route('/setup', methods=['GET', 'POST'])
def coach_setup():
    existing = get_coach_settings()
    if request.method == 'POST':
        naam = request.form.get('naam', '').strip()
        gemeente = request.form.get('gemeente', '').strip()
        uren = request.form.get('uren_per_week', '').strip()
        if not naam:
            flash('Vul uw naam in.', 'danger')
            return render_template('coach_setup.html', gemeenten=GEMEENTEN, existing=existing)
        if not gemeente or gemeente not in GEMEENTEN:
            flash('Selecteer een gemeente.', 'danger')
            return render_template('coach_setup.html', gemeenten=GEMEENTEN, existing=existing)
        try:
            uren_val = float(uren.replace(',', '.'))
            if uren_val <= 0:
                raise ValueError
        except (ValueError, AttributeError):
            flash('Vul het aantal uren per week in (bijv. 8 of 7.5).', 'danger')
            return render_template('coach_setup.html', gemeenten=GEMEENTEN, existing=existing)
        save_coach_settings(naam, gemeente, uren_val)
        flash(f'Welkom, {naam}!', 'success')
        return redirect(url_for('main.index'))
    return render_template('coach_setup.html', gemeenten=GEMEENTEN, existing=existing)


# ── Homepage ──────────────────────────────────────────────────────────────────

@bp.route('/')
def index():
    if not g.coach.get('naam'):
        return redirect(url_for('main.coach_setup'))
    clients = get_all_clients()
    update_status = get_update_status()
    status = {
        c['id']: {
            1: get_vl1(c['id']) is not None,
            2: get_vl2(c['id']) is not None,
            3: get_vl3(c['id']) is not None,
        }
        for c in clients
    }
    active_clients = [c for c in clients if not all(status[c['id']].values())]
    done_clients   = [c for c in clients if all(status[c['id']].values())]
    return render_template('index.html',
        active_clients=active_clients,
        done_clients=done_clients,
        status=status,
        update_status=update_status)


# ── Dashboard ─────────────────────────────────────────────────────────────────

@bp.route('/dashboard')
def dashboard():
    if not g.coach.get('naam'):
        return redirect(url_for('main.coach_setup'))
    periode = request.args.get('periode', 'alles')
    if periode not in ('kwartaal', 'jaar', 'alles'):
        periode = 'alles'
    active_tab = request.args.get('tab', 'caseload')
    if active_tab not in ('caseload', 'impact', 'doorstroom'):
        active_tab = 'caseload'
    status_filter = request.args.get('status', 'alle')
    radar_group = request.args.get('radar_group', 'alle')
    radar_value = request.args.get('radar_value', 'alle')
    data = get_dashboard_data(periode, status=status_filter, radar_group=radar_group, radar_value=radar_value)

    uren_per_dossier = None
    uren_per_week = g.coach.get('uren_per_week')
    if uren_per_week:
        try:
            uren_per_week = float(uren_per_week)
            if uren_per_week > 0 and data['kpis']['actief'] > 0:
                uren_per_dossier = round(uren_per_week / data['kpis']['actief'], 2)
        except (TypeError, ValueError):
            uren_per_week = None

    coach_overview = {
        'uren_per_week': uren_per_week,
        'uren_per_dossier': uren_per_dossier,
        'signalen_totaal': len(data['signalen']),
    }

    return render_template('dashboard.html', periode=periode, coach_overview=coach_overview, active_tab=active_tab, **data)


@bp.route('/update/apply', methods=['POST'])
def update_apply():
    if not g.coach.get('naam'):
        return redirect(url_for('main.coach_setup'))

    try:
        update_result = start_self_update(force=True)
    except UpdateError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('main.index'))

    _schedule_process_exit(delay_seconds=1.2)
    return render_template('update_installing.html', **update_result)


# ── Client toevoegen / verwijderen ────────────────────────────────────────────

@bp.route('/client/add', methods=['GET', 'POST'])
def client_add():
    if request.method == 'POST':
        voornaam = request.form.get('voornaam', '').strip()
        if not voornaam:
            flash('Voornaam mag niet leeg zijn.', 'danger')
            return render_template('client_add.html')

        parts = voornaam.split()
        first_name = parts[0]

        if len(parts) > 1:
            # Naam bevat spatie — controleer of het een achternaam of een initiaal is
            second_stripped = parts[1].rstrip('.')
            if len(second_stripped) > 1 or len(parts) > 2:
                # Ziet eruit als een achternaam → blokkeren
                flash(
                    'Voer alleen de voornaam in, geen achternaam. '
                    'Gebruik een initiaal (bijv. "Anna V.") als er al een cliënt met dezelfde voornaam bestaat.',
                    'danger'
                )
                return render_template('client_add.html', voornaam=voornaam)
            else:
                # "Anna V." formaat — alleen toegestaan als er al een "Anna" bestaat
                existing = get_clients_by_first_name(first_name)
                if not existing:
                    flash(
                        f'Een initiaal is alleen nodig als er al een cliënt met de voornaam "{first_name}" bestaat. '
                        f'Bestaat de naam al? Voer anders alleen de voornaam in.',
                        'warning'
                    )
                    return render_template('client_add.html', voornaam=voornaam)
        else:
            # Alleen voornaam — controleer op exacte duplicaten
            existing = get_clients_by_first_name(first_name)
            exact_duplicates = [c for c in existing if c['voornaam'] == first_name]
            if exact_duplicates:
                flash(
                    f'Er bestaat al een cliënt met de naam "{first_name}". '
                    f'Voeg een initiaal toe om ze te onderscheiden, bijv. "{first_name} A."',
                    'danger'
                )
                return render_template('client_add.html', voornaam=voornaam)

        add_client(voornaam)
        flash(f'{voornaam} is toegevoegd.', 'success')
        return redirect(url_for('main.index'))
    return render_template('client_add.html')


@bp.route('/client/<int:client_id>/delete', methods=['POST'])
def client_delete(client_id):
    client = get_client(client_id)
    if client:
        delete_client(client_id)
        flash(f'{client["voornaam"]} is verwijderd.', 'success')
    return redirect(url_for('main.index'))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_int(val, min_val=0, max_val=10):
    try:
        v = int(val)
        return max(min_val, min(max_val, v))
    except (TypeError, ValueError):
        return None


def _safe_float(val, min_val=0.0, max_val=9999.0):
    try:
        v = round(float(str(val).replace(',', '.')), 1)
        return max(min_val, min(max_val, v))
    except (TypeError, ValueError):
        return None


def _parse_sw(form):
    """Parse sw_q1..sw_q44 from form, return dict."""
    return {f'sw_q{i}': _safe_int(form.get(f'sw_q{i}')) for i in range(1, 45)}


# ── Vragenlijst 1 — Intake ───────────────────────────────────────────────────

@bp.route('/client/<int:client_id>/vragenlijst/1', methods=['GET', 'POST'])
def vragenlijst_1(client_id):
    client = get_client(client_id)
    if not client:
        flash('Cliënt niet gevonden.', 'danger')
        return redirect(url_for('main.index'))

    existing = get_vl1(client_id)

    if request.method == 'POST':
        f = request.form
        data = {
            'verwijzer': f.get('verwijzer', '').strip() or None,
            'huisartsenpraktijk': f.get('huisartsenpraktijk', '').strip() or None,
            'geslacht': f.get('geslacht', '').strip() or None,
            'leeftijdscategorie': f.get('leeftijdscategorie', '').strip() or None,
            'woonsituatie': f.get('woonsituatie', '').strip() or None,
            'werkstatus': f.get('werkstatus', '').strip() or None,
            'eerder_hulp': f.get('eerder_hulp', '').strip() or None,
            'huisartsbezoeken': _safe_int(f.get('huisartsbezoeken'), 0, 999),
            **_parse_sw(f),
        }
        save_vl1(client_id, data)
        flash('Vragenlijst 1 opgeslagen.', 'success')
        return redirect(url_for('main.vragenlijst_1_view', client_id=client_id))

    return render_template('vragenlijst_1.html', client=client, existing=existing,
                           sw_questions=SW_QUESTIONS)


# ── Vragenlijst 2 — Uitstroom ────────────────────────────────────────────────

@bp.route('/client/<int:client_id>/vragenlijst/2', methods=['GET', 'POST'])
def vragenlijst_2(client_id):
    client = get_client(client_id)
    if not client:
        flash('Cliënt niet gevonden.', 'danger')
        return redirect(url_for('main.index'))

    existing = get_vl2(client_id)

    if request.method == 'POST':
        f = request.form
        hoofdreden = f.get('hoofdreden', '').strip()
        uitval_ja_nee = f.get('uitval_ja_nee', '').strip()
        doorverwezen_ja_nee = f.get('doorverwezen_ja_nee', '').strip()

        errors = []

        if not hoofdreden:
            errors.append('Selecteer een hoofdreden aanmelding.')
        if not uitval_ja_nee:
            errors.append('Geef aan of de deelnemer uitgevallen is.')
        if uitval_ja_nee == 'ja' and not f.get('uitval_reden', '').strip():
            errors.append('Selecteer de voornaamste reden van uitval.')
        if doorverwezen_ja_nee == 'ja' and not f.get('doorverwezen_naar', '').strip():
            errors.append('Geef aan waarnaar de deelnemer is doorverwezen.')

        # Datumvolgorde: verwijzing ≤ intake ≤ start activiteit
        def _parse_date(s):
            if not s:
                return None
            try:
                return date.fromisoformat(s)
            except ValueError:
                return None

        d_verw = _parse_date(f.get('datum_verwijzing'))
        d_intake = _parse_date(f.get('datum_intake'))
        d_start = _parse_date(f.get('datum_start_activiteit'))

        if d_verw and d_intake and d_verw > d_intake:
            errors.append('Datum verwijzing moet vóór of gelijk aan datum eerste gesprek (intake) liggen.')
        if d_intake and d_start and d_intake > d_start:
            errors.append('Datum eerste gesprek (intake) moet vóór of gelijk aan datum start activiteit liggen.')
        if d_verw and d_start and not d_intake and d_verw > d_start:
            errors.append('Datum verwijzing moet vóór of gelijk aan datum start activiteit liggen.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('vragenlijst_2.html', client=client, existing=existing)

        data = {
            'hoofdreden': hoofdreden,
            'uitstroom_naar': f.get('uitstroom_naar', '').strip() or None,
            'geslacht': f.get('geslacht', '').strip() or None,
            'leeftijdscategorie': f.get('leeftijdscategorie', '').strip() or None,
            'woonsituatie': f.get('woonsituatie', '').strip() or None,
            'werkstatus': f.get('werkstatus', '').strip() or None,
            'datum_verwijzing': f.get('datum_verwijzing') or None,
            'datum_intake': f.get('datum_intake') or None,
            'datum_start_activiteit': f.get('datum_start_activiteit') or None,
            'contactmomenten_ff': _safe_float(f.get('contactmomenten_ff'), 0.0, 9999.0),
            'contactmomenten_tel': _safe_float(f.get('contactmomenten_tel'), 0.0, 9999.0),
            'uitval_ja_nee': uitval_ja_nee,
            'uitval_reden': (f.get('uitval_reden', '').strip() or None) if uitval_ja_nee == 'ja' else None,
            'doorverwezen_ja_nee': doorverwezen_ja_nee or None,
            'doorverwezen_naar': (f.get('doorverwezen_naar', '').strip() or None) if doorverwezen_ja_nee == 'ja' else None,
            'terugkoppeling': f.get('terugkoppeling', '').strip() or None,
        }
        save_vl2(client_id, data)
        flash('Vragenlijst 2 opgeslagen.', 'success')
        return redirect(url_for('main.vragenlijst_2_view', client_id=client_id))

    return render_template('vragenlijst_2.html', client=client, existing=existing)


# ── Vragenlijst 3 — Opvolging ────────────────────────────────────────────────

@bp.route('/client/<int:client_id>/vragenlijst/3', methods=['GET', 'POST'])
def vragenlijst_3(client_id):
    client = get_client(client_id)
    if not client:
        flash('Cliënt niet gevonden.', 'danger')
        return redirect(url_for('main.index'))

    existing = get_vl3(client_id)

    if request.method == 'POST':
        f = request.form
        continuering = f.get('continuering', '').strip()

        if not continuering:
            flash('Geef aan of de deelnemer nog meedoet aan de activiteit.', 'danger')
            return render_template('vragenlijst_3.html', client=client, existing=existing,
                                   sw_questions=SW_QUESTIONS)

        data = {
            'continuering': continuering,
            'reden_stoppen': (f.get('reden_stoppen', '').strip() or None) if continuering in ('nee_gestopt', 'nee_nooit') else None,
            'behoefte_ondersteuning': f.get('behoefte_ondersteuning', '').strip() or None,
            'tevredenheidscijfer': _safe_int(f.get('tevredenheidscijfer'), 1, 10),
            'huisartsbezoeken': _safe_int(f.get('huisartsbezoeken'), 0, 999),
            **_parse_sw(f),
        }
        save_vl3(client_id, data)
        flash('Vragenlijst 3 opgeslagen.', 'success')
        return redirect(url_for('main.vragenlijst_3_view', client_id=client_id))

    return render_template('vragenlijst_3.html', client=client, existing=existing,
                           sw_questions=SW_QUESTIONS)


# ── View routes (read-only na invullen) ──────────────────────────────────────

@bp.route('/client/<int:client_id>/vragenlijst/1/view')
def vragenlijst_1_view(client_id):
    client = get_client(client_id)
    if not client:
        flash('Cliënt niet gevonden.', 'danger')
        return redirect(url_for('main.index'))
    vl1 = get_vl1(client_id)
    if not vl1:
        return redirect(url_for('main.vragenlijst_1', client_id=client_id))
    scores = calc_sw_scores(vl1)
    scores_list = [scores.get(dim) for dim in SW_QUESTIONS]
    return render_template('vragenlijst_1_view.html', client=client, row=vl1,
                           sw_questions=SW_QUESTIONS, scores=scores,
                           scores_list=scores_list)


@bp.route('/client/<int:client_id>/vragenlijst/2/view')
def vragenlijst_2_view(client_id):
    client = get_client(client_id)
    if not client:
        flash('Cliënt niet gevonden.', 'danger')
        return redirect(url_for('main.index'))
    vl2 = get_vl2(client_id)
    if not vl2:
        return redirect(url_for('main.vragenlijst_2', client_id=client_id))
    return render_template('vragenlijst_2_view.html', client=client, vl2=vl2)


@bp.route('/client/<int:client_id>/vragenlijst/3/view')
def vragenlijst_3_view(client_id):
    client = get_client(client_id)
    if not client:
        flash('Cliënt niet gevonden.', 'danger')
        return redirect(url_for('main.index'))
    vl3 = get_vl3(client_id)
    if not vl3:
        return redirect(url_for('main.vragenlijst_3', client_id=client_id))
    vl1 = get_vl1(client_id)
    scores3 = calc_sw_scores(vl3)
    scores1 = calc_sw_scores(vl1) if vl1 else None
    scores3_list = [scores3.get(dim) for dim in SW_QUESTIONS]
    scores1_list = [scores1.get(dim) if scores1 else None for dim in SW_QUESTIONS]
    return render_template('vragenlijst_3_view.html', client=client, row=vl3,
                           sw_questions=SW_QUESTIONS, scores=scores3, scores1=scores1,
                           scores_list=scores3_list, scores1_list=scores1_list,
                           vl1_row=vl1)


# ── Export naar Excel ────────────────────────────────────────────────────────

@bp.route('/export')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import hashlib
    rows, sw_vl1, sw_vl3 = get_all_for_export()

    # Build deterministic anonymous IDs: 3 uppercase letters + 3 digits per client
    anon_map = {}
    for row in rows:
        cid = row['id']
        h = hashlib.sha256(f"wor-anon-{cid}".encode()).hexdigest()
        letters = ''.join(chr(65 + int(h[i], 16) % 26) for i in range(3))
        digits = ''.join(str(int(h[i + 3], 16) % 10) for i in range(3))
        anon_map[cid] = letters + digits

    # Coach info for export
    coach = get_coach_settings()
    coach_id = coach.get('coach_id', '')
    gemeente = coach.get('gemeente', '')
    uren_per_week = coach.get('uren_per_week', '')

    wb = Workbook()

    # ── Sheet 1: Overzicht ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overzicht"
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")

    headers = [
        'Deelnemer ID', 'Welzijnscoach ID', 'Gemeente', 'Uren beschikbaar voor WoR', 'Aangemaakt op',
        # VL1
        'VL1: Verwijzer', 'VL1: Geslacht', 'VL1: Leeftijdscategorie',
        'VL1: Woonsituatie', 'VL1: Werkstatus', 'VL1: Eerder hulp',
        'VL1: Huisartsbezoeken',
        # VL2
        'VL2: Hoofdreden', 'VL2: Uitstroom naar',
        'VL2: Datum verwijzing', 'VL2: Datum intake', 'VL2: Datum start activiteit',
        'VL2: Contactmomenten f2f', 'VL2: Contactmomenten tel',
        'VL2: Uitval', 'VL2: Uitval reden',
        'VL2: Doorverwezen', 'VL2: Doorverwezen naar',
        'VL2: Terugkoppeling',
        # VL3
        'VL3: Continuering', 'VL3: Reden stoppen',
        'VL3: Behoefte ondersteuning', 'VL3: Tevredenheidscijfer',
        'VL3: Huisartsbezoeken',
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    for row in rows:
        ws1.append([anon_map[row['id']], coach_id, gemeente, uren_per_week, row['aangemaakt_op']]
                   + [row[i] for i in range(3, len(row))])

    for col in ws1.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(w + 2, 35)

    # ── Sheet 2: Spinnenweb VL1 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Spinnenweb Intake (VL1)")
    sw_headers = ['Deelnemer ID'] + [f"Q{n}: {txt}" for dim, qs in SW_QUESTIONS.items() for (n, txt) in qs]
    sw_headers += ['Gem. Lichaamsfuncties', 'Gem. Mentaal welbevinden', 'Gem. Zingeving',
                   'Gem. Kwaliteit van leven', 'Gem. Meedoen', 'Gem. Dagelijks functioneren']
    ws2.append(sw_headers)
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    for row in rows:
        cid = row['id']
        sw = sw_vl1.get(cid)
        if sw:
            scores = calc_sw_scores(sw)
            q_vals = [sw[f'sw_q{i}'] for i in range(1, 45)]
            dim_avgs = [scores.get(d) for d in SW_QUESTIONS]
            ws2.append([anon_map[cid]] + q_vals + dim_avgs)
        else:
            ws2.append([anon_map[cid]] + [''] * 50)

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 12
    ws2.column_dimensions['A'].width = 18

    # ── Sheet 3: Spinnenweb VL3 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Spinnenweb Opvolging (VL3)")
    ws3.append(sw_headers)
    for cell in ws3[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    for row in rows:
        cid = row['id']
        sw = sw_vl3.get(cid)
        if sw:
            scores = calc_sw_scores(sw)
            q_vals = [sw[f'sw_q{i}'] for i in range(1, 45)]
            dim_avgs = [scores.get(d) for d in SW_QUESTIONS]
            ws3.append([anon_map[cid]] + q_vals + dim_avgs)
        else:
            ws3.append([anon_map[cid]] + [''] * 50)

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 12
    ws3.column_dimensions['A'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"wor_export_{date.today().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Heartbeat (browser-open watchdog) ────────────────────────────────────────

@bp.route('/ping', methods=['POST'])
def ping():
    current_app.config['LAST_PING'] = time.time()
    return '', 204


# ── 404 ───────────────────────────────────────────────────────────────────────

@bp.app_errorhandler(404)
def page_not_found(e):
    return render_template('404.html', coach=get_coach_settings(), app_version=APP_VERSION), 404
